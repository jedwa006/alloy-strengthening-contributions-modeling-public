"""Loaders for the user's source workbooks (private — not in public mirror).

Lazy: each loader pulls one workbook on demand. Returns a pandas DataFrame
when pandas is available, otherwise a list of dicts.

Workbook layouts are documented in `data/README.md`. These loaders make
no schema-conformance promises — they exist to bridge from xlsx into the
model's calibration utilities; downstream code should validate fields.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

# Repo-root-relative paths; resolved at import time so missing files raise
# a useful error (ImportError) rather than at first use.
_REPO_ROOT = Path(__file__).resolve().parents[3]

NANOINDENT_WORKBOOK = (
    _REPO_ROOT / "data" / "nanoindentation" / "experimental"
    / "AllSamples M54_HvsErvsThicknessToCrossSection.xlsx"
)

XRD_WORKBOOK_4CR = (
    _REPO_ROOT / "data" / "xrd" / "experimental"
    / "M54_MoXRD_ConvertedToCuSpectrum_Workbook_0p20p40p60p.xlsx"
)


@dataclass(frozen=True)
class NanoIndentRow:
    """One nanoindent point from the AllSamples workbook."""

    cw_pct: float
    rel_x: float  # relative position along scan (depth proxy, 0 = surface)
    H_GPa: float
    H_std_GPa: float
    Er_GPa: float
    Er_std_GPa: float


def load_nanoindentation(cw_pct: float) -> list[NanoIndentRow]:
    """Pull the per-indent (H, Er) rows for one CR condition.

    Sheets in the workbook are named '0%', '20%', '40%', '60%'. Each has
    columns: 'Relative X Position', 'H (GPa) X%', 'stdDev(H) (GPa)',
    'Er (GPa) X%', 'stdDev (Er) (GPa)'. Other columns (aggregates, ASTAR)
    are ignored here.
    """
    if not NANOINDENT_WORKBOOK.exists():
        raise FileNotFoundError(
            f"Nanoindentation workbook not found at {NANOINDENT_WORKBOOK}; "
            "this loader only works in the private repo (file is excluded "
            "from the public mirror per data/README.md)."
        )
    try:
        import openpyxl  # noqa: F401  (used implicitly via load_workbook)
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("load_nanoindentation requires openpyxl") from e

    sheet = f"{int(cw_pct)}%"
    wb = load_workbook(NANOINDENT_WORKBOOK, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise KeyError(f"sheet {sheet!r} not in {wb.sheetnames}")
    ws = wb[sheet]

    rows: list[NanoIndentRow] = []
    for r in ws.iter_rows(values_only=True, min_row=3):
        x, h, h_std, er, er_std = r[0], r[1], r[2], r[3], r[4]
        if x is None or h is None or er is None:
            continue
        try:
            rows.append(
                NanoIndentRow(
                    cw_pct=cw_pct,
                    rel_x=float(x),
                    H_GPa=float(h),
                    H_std_GPa=float(h_std) if h_std is not None else 0.0,
                    Er_GPa=float(er),
                    Er_std_GPa=float(er_std) if er_std is not None else 0.0,
                )
            )
        except (TypeError, ValueError):
            continue
    return rows


@dataclass(frozen=True)
class XRDPoint:
    """One (2θ, intensity) point from the Cu-equivalent spectrum."""

    cw_pct: float
    two_theta_deg: float
    intensity_counts: float
    d_spacing_A: float | None = None


def load_xrd_spectrum(cw_pct: float) -> list[XRDPoint]:
    """Pull the Cu-equivalent XRD spectrum at one CR condition.

    Workbook sheets: 'Cu Data 0p', 'Cu Data 20p', 'Cu Data 40p', 'Cu Data 60p'.
    Columns: No., Pos. [°2θ], Intensity [Counts], Icalc, Iback, ESD,
             D spacings, CT [s].
    """
    if not XRD_WORKBOOK_4CR.exists():
        raise FileNotFoundError(
            f"XRD workbook not found at {XRD_WORKBOOK_4CR}; private file."
        )
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("load_xrd_spectrum requires openpyxl") from e

    sheet = f"Cu Data {int(cw_pct)}p"
    wb = load_workbook(XRD_WORKBOOK_4CR, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise KeyError(f"sheet {sheet!r} not in {wb.sheetnames}")
    ws = wb[sheet]

    pts: list[XRDPoint] = []
    for r in ws.iter_rows(values_only=True, min_row=3):
        if r[0] is None:  # 'No.' column blank → skip
            continue
        try:
            two_theta = float(r[1])
            intensity = float(r[2])
        except (TypeError, ValueError):
            continue
        d = None
        try:
            d = float(r[6]) if r[6] is not None else None
        except (TypeError, ValueError):
            d = None
        pts.append(XRDPoint(cw_pct=cw_pct, two_theta_deg=two_theta,
                            intensity_counts=intensity, d_spacing_A=d))
    return pts
